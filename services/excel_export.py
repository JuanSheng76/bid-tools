"""Excel 导出公共格式。"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADER_FILL = PatternFill("solid", fgColor="2563A8")
ALT_ROW_FILL = PatternFill("solid", fgColor="F5F8FB")
WHITE_FONT = "FFFFFF"
TEXT_COLOR = "182638"
LINE_COLOR = "DCE4ED"


def style_export_sheet(
    worksheet,
    column_widths: dict[str, float],
    number_formats: dict[str, str] | None = None,
) -> None:
    """为后台导出的数据表应用统一、可读的基础格式。"""

    thin_bottom = Border(
        bottom=Side(style="thin", color=LINE_COLOR)
    )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 24

    for cell in worksheet[1]:
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE_FONT)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=2), start=2
    ):
        worksheet.row_dimensions[row_index].height = 21
        for cell in row:
            cell.font = Font(name="Arial", size=10, color=TEXT_COLOR)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_bottom
            if row_index % 2 == 0:
                cell.fill = ALT_ROW_FILL

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    for column, number_format in (number_formats or {}).items():
        for cell in worksheet[column][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = number_format
